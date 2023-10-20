import openpyxl
import numpy as np
from numpy.linalg import inv

# 直接消耗系数算法
def calculate_direct_consumption_coefficients(file_path, sheet_name, start_row, end_row, start_col, end_col, divider_row):
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path)

        # 选择工作表
        worksheet = workbook[sheet_name]

        # 获取分母行的数据，用于除法操作
        divider_values = [worksheet.cell(row=divider_row, column=col).value for col in range(start_col, end_col + 1)]

        # 创建一个空数组来存储直接消耗系数数据
        direct_consumption_coefficient_array = []

        # 循循环遍历指定范围内的单元格
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                divider_value = divider_values[col - start_col]
                if divider_value is not None:
                    result = cell_value / divider_value
                else:
                    result = None
                row_data.append(result)
            direct_consumption_coefficient_array.append(row_data)

        # 创建一个新的工作表并命名为 "直接消耗系数"
        new_sheet = workbook.create_sheet("直接消耗系数")

        # 将结果写入新工作表
        for row_index, row_data in enumerate(direct_consumption_coefficient_array, 1):
            for col_index, value in enumerate(row_data, 1):
                new_sheet.cell(row=row_index, column=col_index, value=value)

        # 保存工作簿
        workbook.save(file_path)
        workbook.close()

        return direct_consumption_coefficient_array
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# 直接分配系数算法
def calculate_direct_allocation_coefficients(file_path, sheet_name, start_row, end_row, start_col, end_col, divider_start_row, total_output_col):
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path)

        # 选择工作表
        worksheet = workbook[sheet_name]

        # 创建一个空数组来存储直接分配系数数据
        direct_allocation_coefficient_array = []

        # 循循环遍历指定范围内的单元格
        for row in range(start_row, end_row + 1):
            row_data = []
            divider_cell = worksheet.cell(row=divider_start_row + row - start_row, column=total_output_col)  # Column 18 corresponds to column 'R'
            divider_value = divider_cell.value
            if divider_value is not None:
                for col in range(start_col, end_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    result = cell_value / divider_value
                    row_data.append(result)
            else:
                row_data = [None] * (end_col - start_col + 1)
            direct_allocation_coefficient_array.append(row_data)

        # 创建一个新的工作表并命名为 "直接分配系数"
        new_sheet = workbook.create_sheet("直接分配系数")

        # 将结果写入新工作表
        for row_index, row_data in enumerate(direct_allocation_coefficient_array, 1):
            for col_index, value in enumerate(row_data, 1):
                new_sheet.cell(row=row_index, column=col_index, value=value)

        # 保存工作簿
        workbook.save(file_path)
        workbook.close()

        return direct_allocation_coefficient_array
    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return None

# 生产活动系数算法
def calculate_production_activity_coefficient(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表 "直接消耗系数"
    worksheet = workbook["直接消耗系数"]

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column

    # 从工作表加载直接消耗系数矩阵
    direct_consumption_matrix = []
    for row in worksheet.iter_rows(values_only=True):
        row_data = []
        for cell_value in row:
            row_data.append(cell_value)
        direct_consumption_matrix.append(row_data)

    # 创建一个单位矩阵与直接消耗系数矩阵相同大小
    identity_matrix = np.identity(num_rows)

    # 执行矩阵减法
    result_matrix = identity_matrix - np.array(direct_consumption_matrix)

    # 创建新的工作表 "生产活动系数"
    new_worksheet = workbook.create_sheet(title="生产活动系数")

    # 将结果矩阵写入新工作表
    for row_index, row in enumerate(result_matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            new_worksheet.cell(row=row_index, column=col_index, value=value)

    # 保存工作簿
    workbook.save(file_path)

    # 关闭工作簿
    workbook.close()

# 完全需要系数算法
def calculate_inverse_matrix(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表 "生产活动系数"
    worksheet = workbook["生产活动系数"]

    # 获取工作表的行数和列数
    num_rows = worksheet.max_row
    num_cols = worksheet.max_column

    # 从工作表加载生产活动系数矩阵
    production_activity_matrix = []
    for row in worksheet.iter_rows(values_only=True):
        row_data = []
        for cell_value in row:
            row_data.append(cell_value)
        production_activity_matrix.append(row_data)

    # 计算矩阵的逆矩阵
    inverse_matrix = inv(np.array(production_activity_matrix))

    # 创建新的工作表 "完全需要系数"
    new_worksheet = workbook.create_sheet(title="完全需要系数")

    # 将逆矩阵写入新工作表
    for row_index, row in enumerate(inverse_matrix, start=1):
        for col_index, value in enumerate(row, start=1):
            new_worksheet.cell(row=row_index, column=col_index, value=value)

    # 保存工作簿
    workbook.save(file_path)

    # 关闭工作簿
    workbook.close()

#完全消耗系数算法
def calculate_consumption_coefficient(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表 "完全需要系数"
    worksheet = workbook["完全需要系数"]

    # 创建新的工作表 "完全消耗系数"
    new_worksheet = workbook.create_sheet(title="完全消耗系数")

    # 遍历原始工作表中的数据，对主对角线的数据减1并写入新工作表，其余数据不变
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        new_row = []
        for col_index, cell_value in enumerate(row, start=1):
            if row_index == col_index:
                new_value = cell_value - 1
            else:
                new_value = cell_value
            new_row.append(new_value)
        new_worksheet.append(new_row)

    # 保存工作簿
    workbook.save(file_path)

    # 关闭工作簿
    workbook.close()

#————————————————————————————————————————————————————————————————————————————————————————————
# 需要编辑的部分
# Excel路径
file_path = r'D:\桌面\中国2015-2020增量表（8部门）新.xlsx'
# Excel中工作表名称
sheet_name = "可比价2020年"

# 中间投入起始行和列
intermediate_input_start_row = 5
intermediate_input_start_col = 4

# 中间投入终止行和列
intermediate_input_end_row = 12
intermediate_input_end_col = 11

# 总投入所在行
total_input = 18

# 总产出所在行和列
total_output_row = 5
total_output_col = 18
#——————————————————————————————————————————————————————————————————————————————————————————————————
# 直接消耗系数
calculate_direct_consumption_coefficients(file_path, sheet_name, intermediate_input_start_row, intermediate_input_end_row, intermediate_input_start_col, intermediate_input_end_col, total_input)

# 直接分配系数
calculate_direct_allocation_coefficients(file_path, sheet_name, intermediate_input_start_row, intermediate_input_end_row, intermediate_input_start_col, intermediate_input_end_col, total_output_row,total_output_col)

# 生产活动系数
calculate_production_activity_coefficient(file_path)

# 完全需要系数
calculate_inverse_matrix(file_path)

#完全消耗系数
calculate_consumption_coefficient(file_path)